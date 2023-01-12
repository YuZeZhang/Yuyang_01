import scipy.io as sio
import pandas as pd
import os
import openpyxl

excel_path = './excel'  # 待合并文件的目录
res_file = 'merge_result'  # 写入的文件名
write_path = './merge_result.xlsx'  # 写入的文件路径

# 遍历每一个待合并的文件
for file in os.listdir(excel_path):
    # 只处理.xlsx格式的文件
    if ('.xlsx' in file):
        # 读取表格文件
        file_path = os.path.join(excel_path, file)
        df = pd.read_excel(file_path)
        # 将数字转为字符串格式，如果表格中没有诸如‘001’的数值可省略
        columns = df.columns
        columns = list(df.columns)
        value = [str] * len(columns)
        str_convert = dict(zip(columns, value))
        df = pd.read_excel(file_path, converters=str_convert)
        # 读取键值，且转为list
        keys = list(df.keys())
        values = df.values.tolist()
        # sheet的名称，默认和待合并文件的文件名一样
        sheet_name = file.replace('.xlsx', '')

        # 开始写入
        # 如果已存在写入文件则加载，否则新建
        if (os.path.exists(write_path)):
            wb = openpyxl.load_workbook(write_path)
        else:
            wb = openpyxl.Workbook()

        # 如果已写入同名sheet则删除（确保写入的内容是最新的）
        if (sheet_name in wb.sheetnames):
            wb.remove_sheet(wb[sheet_name])
        # 创建sheet,title属性为sheet的名称
        wb.create_sheet(title=sheet_name, index=0)
        # 按行写入键值
        wb[sheet_name].append(keys)
        for value in values:
            wb[sheet_name].append(value)
        # 清除空表
        for name in wb.sheetnames:
            ws = wb[name]
            if (ws.cell(1, 1).value == None):
                wb.remove_sheet(ws)
        # 保存
        wb.save(write_path)




